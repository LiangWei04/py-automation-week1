"""
Path: src/csv_to_excel_report.py

Purpose: Read multiple CSVs, combine, compute revenue, produce a styled Excel report with:

Sheet Transactions (cleaned rows + computed revenue = units*unit_price)

Sheet Summary (pivot by region and product: sum units and revenue)

A bar chart on Summary showing revenue by region

Basic formatting (bold headers, thousands separators, currency for revenue)

CLI spec (must work exactly):

python src/csv_to_excel_report.py \
  --inputs data/sales_q1.csv data/sales_q2.csv data/sales_q3.csv data/sales_q4.csv \
  --out reports/sales_report.xlsx


Acceptance criteria:

Creates reports/sales_report.xlsx

Transactions has all rows from inputs + revenue column (numeric)

Summary pivot exact columns: region, product, units_sum, revenue_sum

A bar chart titled “Revenue by Region” using revenue_sum aggregated by region

Number formats: units_sum as integer; revenue/revenue_sum as currency

No exceptions on missing optional columns (but fail if required headers missing)

Implementation notes (pseudocode-level, exact ops to perform):

Read CSVs with pd.read_csv, pd.concat (ignore_index=True)

Validate required cols: {"date","region","product","units","unit_price"}

Cast units to Int64, unit_price to float, compute revenue = units*unit_price

Build pivot with groupby(["region","product"]).agg({"units":"sum","revenue":"sum"})

Write with pd.ExcelWriter(engine="openpyxl")

Use openpyxl to:

bold header rows,

apply number formats: #,##0 for units; "$"#,##0.00 for revenue,

insert BarChart on Summary (X = regions, Y = revenue_sum)
"""
# csv_to_excel_report.py
import argparse
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

LIGHT_GRAY = PatternFill("solid", fgColor="DDDDDD")
BOLD = Font(bold=True)
THIN_BORDER = Border(top=Side(style="thin"))
CENTER = Alignment(horizontal="center")

def arg_parse():
    p = argparse.ArgumentParser(
        description="Combine CSVs, compute revenue, and write an Excel report."
    )
    p.add_argument("--inputs", nargs="+", required=True,
                   help="One or more input CSVs (e.g., --inputs data/q1.csv data/q2.csv)")
    p.add_argument("--out", required=True,
                   help="Output Excel file (e.g., reports/sales_report.xlsx)")
    return p.parse_args()

def _autofit(ws, start_col=1, end_col=None, start_row=1, end_row=None, padding=2):
    if end_col is None:
        end_col = ws.max_column
    if end_row is None:
        end_row = ws.max_row
    for col_idx in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(start_row, end_row + 1):
            v = ws[f"{col_letter}{row}"].value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = max(10, max_len + padding)

def write_summary_sheet(writer: pd.ExcelWriter, df: pd.DataFrame):
    """
    Assumes df has: date, region, product, units, unit_price, revenue
    Creates 'Summary' to match your checklist.
    """

    # ---- Aggregations ----
    pivot = (df.groupby(["region", "product"], as_index=False)
               .agg(units_sum=("units", "sum"), revenue_sum=("revenue", "sum"))
               .sort_values(["region", "product"], kind="stable"))

    region_totals = (df.groupby("region", as_index=False)
                       .agg(revenue_sum=("revenue", "sum"))
                       .sort_values("revenue_sum", ascending=False, kind="stable"))

    product_totals = (df.groupby("product", as_index=False)
                        .agg(revenue_sum=("revenue", "sum"))
                        .sort_values("revenue_sum", ascending=False, kind="stable"))

    # Monthly trend (optional/bonus)
    if pd.api.types.is_datetime64_any_dtype(df["date"]):
        month = df["date"].dt.strftime("%Y-%m")
    else:
        month = df["date"].astype(str).str[:7]
    monthly = (df.assign(month=month)
                 .groupby("month", as_index=False)
                 .agg(units_sum=("units", "sum"), revenue_sum=("revenue", "sum"))
                 .sort_values("month", kind="stable"))

    # ---- KPI values ----
    total_revenue = float(df["revenue"].sum())
    total_units = int(df["units"].sum())
    aov = (total_revenue / total_units) if total_units else 0.0
    n_transactions = int(len(df))
    top_region = region_totals.iloc[0]["region"] if not region_totals.empty else ""
    top_product = product_totals.iloc[0]["product"] if not product_totals.empty else ""

    # ---- Create & format sheet ----
    pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)
    ws = writer.book["Summary"]
    ws.column_dimensions.width = 19
    # 1) KPI block A1:D6 (we use A for labels, B for values; C:D left blank but reserved)
    ws["A1"].value = "Label"; ws["B1"].value = "Value"
    ws["A1"].font = BOLD; ws["B1"].font = BOLD
    ws["A1"].fill = LIGHT_GRAY; ws["B1"].fill = LIGHT_GRAY

    kpi_rows = [
        ("Total Revenue", total_revenue, "$#,##0.00"),
        ("Total Units", total_units, "#,##0"),
        ("Avg Order Value (AOV)", aov, "$#,##0.00"),
        ("# Transactions", n_transactions, "#,##0"),
        ("Top Region (by Rev)", top_region, None),
        ("Top Product (by Rev)", top_product, None),
    ]
    for i, (label, value, fmt) in enumerate(kpi_rows, start=2):
        ws[f"A{i}"].value = label
        ws[f"B{i}"].value = value
        if fmt:
            ws[f"B{i}"].number_format = fmt

    # 2) Region × Product pivot at A8 with exact headers
    start_row = 8
    headers = ["region", "product", "units_sum", "revenue_sum"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=j, value=h)
        cell.font = BOLD; cell.fill = LIGHT_GRAY; cell.alignment = CENTER

    r = start_row + 1
    for _, row in pivot.iterrows():
        ws.cell(row=r, column=1, value=row["region"])
        ws.cell(row=r, column=2, value=row["product"])
        cu = ws.cell(row=r, column=3, value=int(row["units_sum"]))
        cr = ws.cell(row=r, column=4, value=float(row["revenue_sum"]))
        cu.number_format = "#,##0"
        cr.number_format = "$#,##0.00"
        r += 1

    last_pivot_row = r - 1
    # Grand total row
    ws.cell(row=r, column=1, value="Total").font = BOLD
    ws.cell(row=r, column=2, value="")
    units_col = get_column_letter(3)
    rev_col = get_column_letter(4)
    ws.cell(row=r, column=3,
            value=f"=SUM({units_col}{start_row+1}:{units_col}{last_pivot_row})").number_format = "#,##0"
    tot_rev_cell = ws.cell(row=r, column=4,
            value=f"=SUM({rev_col}{start_row+1}:{rev_col}{last_pivot_row})")
    tot_rev_cell.number_format = "$#,##0.00"
    for c in range(1, 5):
        ws.cell(row=r, column=c).font = BOLD
        ws.cell(row=r, column=c).border = THIN_BORDER

    after_pivot_row = r + 3  # spacing before monthly table

    # 3) Region totals at G8 (G:H)
    rt_col = 7  # G
    ws.cell(row=start_row, column=rt_col, value="region").font = BOLD
    ws.cell(row=start_row, column=rt_col, value="region").fill = LIGHT_GRAY
    ws.cell(row=start_row, column=rt_col+1, value="revenue_sum").font = BOLD
    ws.cell(row=start_row, column=rt_col+1, value="revenue_sum").fill = LIGHT_GRAY

    rr = start_row + 1
    for _, row in region_totals.iterrows():
        ws.cell(row=rr, column=rt_col, value=row["region"])
        c = ws.cell(row=rr, column=rt_col+1, value=float(row["revenue_sum"]))
        c.number_format = "$#,##0.00"
        rr += 1

    # 4) Product totals at J8 (J:K)
    pt_col = 10  # J
    ws.cell(row=start_row, column=pt_col, value="product").font = BOLD
    ws.cell(row=start_row, column=pt_col, value="product").fill = LIGHT_GRAY
    ws.cell(row=start_row, column=pt_col+1, value="revenue_sum").font = BOLD
    ws.cell(row=start_row, column=pt_col+1, value="revenue_sum").fill = LIGHT_GRAY

    pr = start_row + 1
    for _, row in product_totals.iterrows():
        ws.cell(row=pr, column=pt_col, value=row["product"])
        c = ws.cell(row=pr, column=pt_col+1, value=float(row["revenue_sum"]))
        c.number_format = "$#,##0.00"
        pr += 1

    # 5) Optional monthly trend starting after pivot + 3
    m_headers = ["month (YYYY-MM)", "units_sum", "revenue_sum"]
    mrow = after_pivot_row
    for j, h in enumerate(m_headers, start=1):
        cell = ws.cell(row=mrow, column=j, value=h)
        cell.font = BOLD; cell.fill = LIGHT_GRAY; cell.alignment = CENTER
    mrow += 1
    for _, row in monthly.iterrows():
        ws.cell(row=mrow, column=1, value=row["month"])
        mu = ws.cell(row=mrow, column=2, value=int(row["units_sum"]))
        mr = ws.cell(row=mrow, column=3, value=float(row["revenue_sum"]))
        mu.number_format = "#,##0"
        mr.number_format = "$#,##0.00"
        mrow += 1

    # 6) Styling & readability
    _autofit(ws)
  # rr currently points to the next empty row after writing region_totals
    rt_header_row = start_row          # G8 header row
    rt_first_data = start_row + 1      # G9 first data row
    rt_last_data  = rr - 1             # last data row actually filled
    
    if rt_last_data >= rt_first_data:
        chart = BarChart()
        chart.title = "Revenue by Region"
        chart.y_axis.title = "Revenue"
        chart.x_axis.title = "Region"
    
        # Categories = region names (column G), Data = revenue_sum (column H)
        cats = Reference(ws, min_col=7, min_row=rt_first_data, max_row=rt_last_data)   # G
        data = Reference(ws, min_col=8, min_row=rt_header_row, max_row=rt_last_data)   # H (includes header)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
    
        chart.height = 12  # tweak size as you like
        chart.width  = 20
        ws.add_chart(chart, "M8")  # place chart at M8 so it doesn’t overlap tables

def main():
    REQUIRED = {"date", "region", "product", "units", "unit_price"}
    args = arg_parse()
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Read & normalize
    frames = []
    for f in args.inputs:
        df_part = pd.read_csv(f)
        df_part.columns = [c.strip().lower() for c in df_part.columns]
        # Try parse date
        if "date" in df_part.columns:
            df_part["date"] = pd.to_datetime(df_part["date"], errors="coerce")
        frames.append(df_part)

    df = pd.concat(frames, ignore_index=True)
    missing = REQUIRED - set(df.columns)
    if missing:
        raise ValueError(f"Missing required column(s): {sorted(missing)}")

    # Compute revenue (ensure numeric)
    df["units"] = pd.to_numeric(df["units"], errors="coerce").fillna(0)
    df["unit_price"] = pd.to_numeric(df["unit_price"], errors="coerce").fillna(0.0)
    df["revenue"] = df["units"] * df["unit_price"]

    # Write Excel
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Transactions sheet
        tx = df.copy()
        # nice column order if present
        cols = [c for c in ["date","region","product","units","unit_price","revenue"] if c in tx.columns]
        tx = tx[cols]
        tx.to_excel(writer, sheet_name="Transactions", index=False)
        # Apply date format + widen column A on Transactions
        ws_tx = writer.book["Transactions"]
        # If your date column is column A:
        for r in range(2, ws_tx.max_row + 1):
            ws_tx[f"A{r}"].number_format = "yyyy-mm-dd"  # or "dd-mmm-yyyy"
        ws_tx.column_dimensions[get_column_letter(1)].width = 14  # widen A
            # Summary sheet
        write_summary_sheet(writer, df)

    print(f"Report written to {out_path}")

if __name__ == "__main__":
    main()
